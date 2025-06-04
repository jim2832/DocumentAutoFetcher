from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import time
import re
import os
from datetime import datetime, timedelta
import holidays
import pandas as pd
from openpyxl import load_workbook
from calendar import monthrange

# 台灣國定假日
tw_holidays = holidays.TW()

def roc_to_ad(roc_date_str):
    year, month, day = map(int, roc_date_str.split('/'))
    return datetime(year + 1911, month, day)

def working_days_diff(start_date, end_date):
    day_count = 0
    current = start_date
    while current <= end_date:
        if current.weekday() < 5 and current not in tw_holidays:
            day_count += 1
        current += timedelta(days=1)
    return day_count - 1

# 初始化瀏覽器
options = Options()
options.add_argument("--start-maximized")
driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 10)
main_eip_window_handle = None # To store the main EIP portal window

# 登入 EIP
def login():
    global main_eip_window_handle
    driver.get("https://eip.taichung.gov.tw/myPortal.do")
    main_eip_window_handle = driver.current_window_handle # Store after initial get
    wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/h6/a'))).click()
    wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[1]/div[2]/form/ul[1]/li[1]/input'))).send_keys("ahya0201")
    wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[1]/div[2]/form/ul[1]/li[2]/input'))).send_keys("gkzBBca0440@")
    input("🔐 登入完成後請按 Enter 繼續...")

# 進入公文系統
def open_odis():
    global main_eip_window_handle
    if driver.current_window_handle != main_eip_window_handle and main_eip_window_handle is not None:
        driver.switch_to.window(main_eip_window_handle)
    else:
        driver.get("https://eip.taichung.gov.tw/myPortal.do")
        main_eip_window_handle = driver.current_window_handle

    gongwen_button = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[2]/div[1]/div[2]/div/div[1]/div/div[2]/div/ul/li[5]/a/div[1]')))
    ActionChains(driver).move_to_element(gongwen_button).click().perform()
    
    odis_window_handle = None
    WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))

    for handle in driver.window_handles:
        if handle != main_eip_window_handle:
            odis_window_handle = handle
            break
    
    if odis_window_handle:
        driver.switch_to.window(odis_window_handle)
        driver.get("https://odis.taichung.gov.tw/HOME/Home_Splitter.aspx")
        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/div[3]/div[1]/table/tbody/tr/td/table/tbody/tr[1]/td[3]/table'))).click()
    else:
        print("❌ 無法切換到公文系統(ODIS)視窗")

# 查詢與分析單月資料
def process_month(start_date_str_roc, end_date_str_roc, odis_window_handle_param):
    output = []
    
    if driver.current_window_handle != odis_window_handle_param:
        driver.switch_to.window(odis_window_handle_param)
        if "odis.taichung.gov.tw" not in driver.current_url:
             driver.get("https://odis.taichung.gov.tw/HOME/Home_Splitter.aspx")
             wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/div[3]/div[1]/table/tbody/tr/td/table/tbody/tr[1]/td[3]/table'))).click()

    driver.switch_to.default_content()
    try:
        WebDriverWait(driver, 10).until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "main")))
    except Exception as e:
        print(f"❌ 無法切換到 main iframe: {e}")
        return

    try:
        wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/form/table[3]/tbody/tr/td/table/tbody/tr[3]/td[2]/span[1]/input"))).clear()
        driver.find_element(By.XPATH, "/html/body/form/table[3]/tbody/tr/td/table/tbody/tr[3]/td[2]/span[1]/input").send_keys(start_date_str_roc)
        driver.find_element(By.XPATH, "/html/body/form/table[3]/tbody/tr/td/table/tbody/tr[3]/td[2]/span[3]/input").clear()
        driver.find_element(By.XPATH, "/html/body/form/table[3]/tbody/tr/td/table/tbody/tr[3]/td[2]/span[3]/input").send_keys(end_date_str_roc)
        driver.find_element(By.XPATH, '//label[contains(text(), "存查")]').click()
        driver.find_element(By.XPATH, "/html/body/form/table[4]/tbody/tr/td/input[3]").click()
        print(f"✅ 查詢 {start_date_str_roc} ~ {end_date_str_roc}")
    except Exception as e:
        print(f"❌ 查詢欄位載入或操作錯誤：{e}")
        return
    
    try:
        select_element = wait.until(EC.presence_of_element_located((
            By.XPATH, "/html/body/form/table[2]/tbody/tr/td/table/tbody/tr/td[2]/select" 
        )))
        select = Select(select_element)
        select.select_by_value("100")
        print("✅ 成功選擇每頁顯示 100 筆資料")
        time.sleep(1)
    except Exception as e:
        print(f"❌ 找不到或無法操作下拉選單 (每頁筆數): {e}")

    try:
        driver.switch_to.default_content()
        driver.switch_to.frame("main")    
        icon_sort_header = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//input[@type='image' and contains(@id, 'img_SignHead')]"))
        )
        icon_sort_header.click()
        print("✅ 成功點擊排序 icon (第一次)")
        time.sleep(0.5)

        icon_sort_header = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//input[@type='image' and contains(@id, 'img_SignHead')]"))
        )
        icon_sort_header.click()
        print("✅ 第二次點擊排序 icon")
        time.sleep(0.5)
    except Exception as e:
        print(f"❌ 點擊排序 icon 失敗：{e}")

    # Initialization for original pagination logic
    page_number = 0 # 0-indexed for current page set
    dec = 0         # For tracking groups of 10 pages (decade)
    # cur_page_number was used in original to count total items clicked, not directly for pagination logic itself.
    # Let's add it if you need to track total items processed this way.
    # items_processed_count = 0 

    while True:
        driver.switch_to.default_content()
        driver.switch_to.frame("main")
        
        file_icons = driver.find_elements(By.XPATH, '//a[img[contains(@src, "../IMAGE/GDOCSIGN_1.gif")]]')
        
        # Check if file_icons is empty on the first pass (page_number==0 and dec==0)
        # This check was not explicitly in original, but useful to add
        if not file_icons and page_number == 0 and dec == 0:
            print(f"⚠️ 在 {start_date_str_roc} ~ {end_date_str_roc} 查詢期間，第一頁無電子檔圖示或查詢無結果。")
            break 
        elif not file_icons: # No files on a subsequent page
            current_page_display = page_number + 1 + (dec * 10)
            print(f"⚠️ 第 {current_page_display} 頁無電子檔圖示可處理。")
            break

        for i, a_tag in enumerate(file_icons):
            current_odis_window = driver.current_window_handle 
            existing_windows_before_click = set(driver.window_handles)
            
            a_tag.click()
            # items_processed_count +=1 # If you want to count items like original cur_page_number
            
            doc_detail_window_handle = None
            try:
                WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(len(existing_windows_before_click) + 1))
                
                for handle_in_loop in driver.window_handles:
                    if handle_in_loop not in existing_windows_before_click:
                        doc_detail_window_handle = handle_in_loop
                        break
                
                if doc_detail_window_handle is None:
                    print(f"❌ 嚴重錯誤：雖然視窗數量增加，但無法識別文件詳細資料的新視窗句柄。跳過此文件。")
                    if driver.current_window_handle != current_odis_window:
                        driver.switch_to.window(current_odis_window)
                    driver.switch_to.default_content()
                    wait.until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "main")))
                    continue 

                driver.switch_to.window(doc_detail_window_handle)
                
                wait.until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "GDOCDetailTopMenu")))
                driver.find_element(By.XPATH, "//td[@class='MenuChild' and contains(@url, 'TableTitle=流程')]").click()

                driver.switch_to.default_content()
                wait.until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "GDOCDetailMain")))

                data_cells = driver.find_elements(By.CSS_SELECTOR, 'td.cssData')
                doc_number = data_cells[0].text.strip() if len(data_cells) > 0 else "無文號資料"
                unit = data_cells[2].text.strip() if len(data_cells) > 2 else "無單位資料"
                person = data_cells[3].text.strip() if len(data_cells) > 3 else "無人員資料"


                decision_date, archive_date, sign_date = None, None, None
                flow_rows = driver.find_elements(By.XPATH, "//tr[@class='cssGridItem' or @class='cssGridAlter']")
                for row in flow_rows:
                    cells = row.find_elements(By.TAG_NAME, "td")
                    if len(cells) >= 6:
                        date_cell_text_parts = cells[0].text.strip().split()
                        
                        if not date_cell_text_parts:
                            print(f"⚠️ 警告：文件 '{doc_number}' 的流程中日期欄位為空。跳過此流程日期處理。")
                            continue 
                        
                        process_date_str_roc_flow = date_cell_text_parts[0] # Renamed to avoid confusion
                        process_type = cells[4].text.strip()

                        try:
                            parsed_date_ad = roc_to_ad(process_date_str_roc_flow)
                            if process_type == "決行(存查)":
                                decision_date = parsed_date_ad
                            elif process_type == "存查":
                                archive_date = parsed_date_ad
                            elif process_type == "簽收":
                                sign_date = parsed_date_ad
                        except ValueError as ve:
                            print(f"⚠️ 警告：無法解析文件 '{doc_number}' 的日期 '{process_date_str_roc_flow}'。錯誤：{ve}。")
                        except Exception as e_date_conv:
                            print(f"⚠️ 警告：轉換日期時發生未知錯誤 '{process_date_str_roc_flow}' for doc '{doc_number}'. Error: {e_date_conv}")

                if decision_date and archive_date:
                    diff = working_days_diff(decision_date, archive_date)
                    if diff > 5:
                        output.append({
                            "公文文號": doc_number,
                            "主辦單位": unit,
                            "主辦人員": person,
                            "決行(存查)日期": decision_date.strftime("%Y/%m/%d"),
                            "存查日期": archive_date.strftime("%Y/%m/%d"),
                            "簽收日期": sign_date.strftime("%Y/%m/%d"),
                            "資料完整性": "✅ 完整 (符合條件)"
                        })
                        print(f"📌 {doc_number} 記錄完成 (天數差異 {diff} > 5)")

            except Exception as e_doc_processing:
                print(f"❌ 處理文件明細時發生錯誤：{e_doc_processing}")
            finally:
                current_handles_after_doc = set(driver.window_handles)
                for handle_to_close in current_handles_after_doc:
                    if handle_to_close not in existing_windows_before_click and handle_to_close != current_odis_window :
                        try:
                            driver.switch_to.window(handle_to_close)
                            driver.close()
                        except Exception as e_closing_doc_window:
                            print(f"注意：嘗試關閉文件詳細視窗時發生錯誤: {e_closing_doc_window}")
                
                driver.switch_to.window(current_odis_window)
                driver.switch_to.default_content()
                try:
                    wait.until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "main")))
                except Exception as e_main_frame_return:
                    print(f"❌ 嚴重錯誤：返回主列表 iframe 時失敗: {e_main_frame_return}。")

        # --- Pagination Logic (Reverted to original) ---
        driver.switch_to.default_content() # Ensure out of any sub-frames
        driver.switch_to.frame("main")    # Enter main frame where pagination exists

        try:
            page_links = wait.until(EC.presence_of_all_elements_located(
                (By.XPATH, "//a[contains(@href, '__doPostBack') and contains(text(), '[')]")
            ))

            page_numbers_elements = [ # Renamed to avoid conflict with 'page_number' variable
                link for link in page_links
                if re.fullmatch(r"\[\s*\d+\s*\]", link.text.strip())
            ]

            dots_elements = [] # Initialize to empty list
            try:
                # find_elements returns empty list if not found, no need for separate try-except for this
                dots_elements = driver.find_elements(By.XPATH, "//a[contains(@href, '__doPostBack') and contains(text(), '[ ... ]')]")
            except Exception as e_find_dots: # Should not happen for find_elements unless XPath is invalid
                print(f"❌ 尋找 [ ... ] 按鈕時發生 XPath 錯誤: {e_find_dots}")


            print(f"目前頁碼 (邏輯)：組 {dec + 1}，頁索引 {page_number} (顯示 {page_number + 1 + dec * 10})")

            if page_number == 9: # Index 9 means it's the 10th page in the current set
                if dots_elements:
                    dots_elements[-1].click() # Click the last found '...' button
                    print("✅ 點擊 [ ... ] 展開更多頁碼")
                    page_number = 0 # Reset page_number for the new set of pages
                    dec += 1        # Increment 'dec' (decade/group counter)
                    time.sleep(1) # Allow page to load
                else:
                    print("⚠️ 沒有更多 [ ... ] 按鈕，可能已是最後一頁或無後續頁組。")
                    break # Break the while True loop for pagination
            else:
                if page_number >= len(page_numbers_elements):
                    print("⚠️ 頁碼索引超出範圍 (目前頁碼元素數量不足)，或已無下一頁可點擊。停止換頁。")
                    break 
                
                next_page_element = page_numbers_elements[page_number]
                print(f"➡️ 換頁到：{next_page_element.text.strip()}")
                next_page_element.click()
                page_number += 1 # Increment page_number for the next page in the current set
                time.sleep(1) # Allow page to load
        
        except Exception as e_pagination: # Catch errors during finding links or clicking
            print(f"❌ 換頁過程中發生錯誤 ({type(e_pagination).__name__})，可能已到最後一頁: {e_pagination}")
            break # Stop processing this month if pagination fails
    # --- End of While True loop for pagination ---

    if output:
        df = pd.DataFrame(output)
        file_path = "存查結果.xlsx"
        if not os.path.exists(file_path):
            df.to_excel(file_path, index=False, engine='openpyxl')
        else:
            try:
                # Load existing workbook to find the last row
                book = load_workbook(file_path)
                sheet_name = book.sheetnames[0] 
                start_row_index = book[sheet_name].max_row
                
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    # Make sure that the writer.book is correctly set by openpyxl when mode='a'
                    # For 'overlay', openpyxl should handle this. We need to ensure header is not written.
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row_index)
                print(f"✅ 寫入 Excel 完成：新增 {len(output)} 筆資料到 {file_path}")
            except Exception as e_excel:
                 print(f"❌ 寫入 Excel 檔案 '{file_path}' 失敗: {e_excel}")
                 timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                 fallback_path = f"存查結果_fallback_{timestamp}.xlsx"
                 df.to_excel(fallback_path, index=False, engine='openpyxl')
                 print(f"⚠️ 已將本次結果儲存到新的 fallback 檔案: {fallback_path}")
    else:
        print(f"⚠️ 月份 {start_date_str_roc} ~ {end_date_str_roc} 無需寫入資料")

# ========== 主流程 ==========
login() 

open_odis() 
odis_system_window_handle = None
if main_eip_window_handle:
    for handle in driver.window_handles:
        if handle != main_eip_window_handle:
            odis_system_window_handle = handle
            break

if not odis_system_window_handle:
    print("❌ 嚴重錯誤：無法獲取公文系統(ODIS)的視窗句柄。程式中止。")
    if driver: driver.quit()
    exit()

print(f"公文系統(ODIS) 視窗句柄: {odis_system_window_handle}")

for month_num in range(1, 13):
    ad_year_for_roc_111 = 111 + 1911 
    start_day_in_month = 1
    num_days_in_month = monthrange(ad_year_for_roc_111, month_num)[1]
    end_day_in_month = num_days_in_month

    roc_start_date_param = f"111/{month_num:02d}/20"
    roc_end_date_param = f"111/{month_num:02d}/{monthrange(ad_year_for_roc_111, month_num)[1]:02d}"

    print(f"\n🔍 處理月份 {month_num} (民國111年): {roc_start_date_param} ~ {roc_end_date_param}")
    
    if driver.current_window_handle != odis_system_window_handle:
        driver.switch_to.window(odis_system_window_handle)
        if "odis.taichung.gov.tw" not in driver.current_url:
             print(f"警告：目前不在預期的ODIS URL。嘗試導航...")
             driver.get("https://odis.taichung.gov.tw/HOME/Home_Splitter.aspx")
             try:
                wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/div[3]/div[1]/table/tbody/tr/td/table/tbody/tr[1]/td[3]/table'))).click()
             except Exception as e_nav:
                print(f"錯誤：在月份循環中導航到ODIS起始頁面失敗: {e_nav}")
                continue 
    process_month(roc_start_date_param, roc_end_date_param, odis_system_window_handle)

input("\n📁 所有月份處理完成，按 Enter 結束程式...")
if driver:
    driver.quit()