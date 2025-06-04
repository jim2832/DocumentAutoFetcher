from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException
import time, re, os
from datetime import datetime, timedelta
import holidays
import pandas as pd
from openpyxl import load_workbook
from calendar import monthrange
import tkinter as tk
from tkinter import simpledialog
from tkinter import messagebox

# 台灣國定假日
tw_holidays = holidays.TW()

def roc_to_ad(roc_date_str):
    year, month, day = map(int, roc_date_str.split('/'))
    return datetime(year + 1911, month, day)

def working_days_diff(start_date, end_date):
    day_count = 0
    current = start_date
    while current <= end_date:
        if current.weekday() < 5 and current not in tw_holidays:
            day_count += 1
        current += timedelta(days=1)
    return day_count - 1

# ========== 彈窗輸入日期 ==========
def get_user_input():
    def on_submit():
        s = start_entry.get()
        e = end_entry.get()
        if not s or not e:
            messagebox.showerror("錯誤", "開始日期和結束日期都必須輸入！")
        else:
            nonlocal start_date, end_date
            start_date, end_date = s, e
            input_win.destroy()

    start_date, end_date = None, None

    root = tk.Tk()
    root.withdraw()  # 隱藏主視窗

    input_win = tk.Toplevel()
    input_win.title("輸入日期")
    input_win.geometry("300x150")
    input_win.resizable(False, False)

    # 開始日期
    tk.Label(input_win, text="開始日期(例如:111/01/01)").pack(pady=(10, 0))
    start_entry = tk.Entry(input_win)
    start_entry.pack()

    # 結束日期
    tk.Label(input_win, text="結束日期(例如:111/12/31)").pack(pady=(10, 0))
    end_entry = tk.Entry(input_win)
    end_entry.pack()

    # 確認按鈕
    tk.Button(input_win, text="確定", command=on_submit).pack(pady=10)

    input_win.grab_set()  # 鎖定彈窗
    root.wait_window(input_win)  # 等待輸入完畢

    return start_date, end_date

# 初始化瀏覽器
options = Options()
options.add_argument("--start-maximized")
driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 10)

# 登入 EIP
def login():
    driver.get("https://eip.taichung.gov.tw/myPortal.do")
    wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/h6/a'))).click()
    wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[1]/div[2]/form/ul[1]/li[1]/input'))).send_keys("ahya0201")
    wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[1]/div[2]/form/ul[1]/li[2]/input'))).send_keys("gkzBBca0440@")

    time.sleep(10)  # 等待輸入完成

    # 等待 /html/body/div[1]/div[1]/div[2]/form/ul[1]/li[4]/input 被按下去
    try:
        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[1]/div[2]/form/ul[1]/li[4]/input'))).click()
        print("✅ 登入成功")
    except TimeoutException:
        print("❌ 登入按鈕未能點擊，請檢查網頁結構或等待時間是否足夠")
        driver.quit()
        exit(1)
    

# 進入公文系統
def open_odis():
    driver.get("https://eip.taichung.gov.tw/myPortal.do")
    gongwen_button = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[2]/div[1]/div[2]/div/div[1]/div/div[2]/div/ul/li[5]/a/div[1]')))
    ActionChains(driver).move_to_element(gongwen_button).click().perform()
    time.sleep(1)

    for handle in driver.window_handles:
        if handle != original_window:
            driver.switch_to.window(handle)
            break
    
    driver.get("https://odis.taichung.gov.tw/HOME/Home_Splitter.aspx")
    wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/div[3]/div[1]/table/tbody/tr/td/table/tbody/tr[1]/td[3]/table'))).click()

# 查詢與分析單月資料
def process_month(start_date, end_date):
    output = []
    driver.switch_to.default_content()

    try:
        WebDriverWait(driver, 10).until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "main")))
    except:
        print("❌ 無法切換到 main iframe")
        return
    
    # 設定查詢區間與勾選「存查」
    try:
        wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/form/table[3]/tbody/tr/td/table/tbody/tr[3]/td[2]/span[1]/input"))).clear()
        driver.find_element(By.XPATH, "/html/body/form/table[3]/tbody/tr/td/table/tbody/tr[3]/td[2]/span[1]/input").send_keys(start_date)
        driver.find_element(By.XPATH, "/html/body/form/table[3]/tbody/tr/td/table/tbody/tr[3]/td[2]/span[3]/input").clear()
        driver.find_element(By.XPATH, "/html/body/form/table[3]/tbody/tr/td/table/tbody/tr[3]/td[2]/span[3]/input").send_keys(end_date)
        driver.find_element(By.XPATH, '//label[contains(text(), "存查")]').click()
        driver.find_element(By.XPATH, "/html/body/form/table[4]/tbody/tr/td/input[3]").click()
        print(f"✅ 查詢 {start_date} ~ {end_date}")
    except Exception as e:
        print(f"❌ 查詢欄位載入錯誤：{e}")
        return
    
    try:
        select_element = wait.until(EC.presence_of_element_located((
            By.XPATH, "/html/body/form/table[2]/tbody/tr/td/table/tbody/tr/td[2]/select" 
        )))
        select = Select(select_element)
        select.select_by_value("100")
        print("✅ 成功選擇 100 筆資料")
        time.sleep(1)
    except:
        print("❌ 找不到下拉選單 select")
    
    frames = driver.find_elements(By.TAG_NAME, "frame")
    print("目前頁面所有 frame：")
    for i, frame in enumerate(frames):
        print(f"Frame {i}: {frame.get_attribute('name')}")

    # STEP 6：排序後找到電子檔(點兩下icon)
    try:
        driver.switch_to.default_content()
        driver.switch_to.frame("main")
        icon = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//input[@type='image' and contains(@id, 'img_SignHead')]"))
        )
        icon.click()
        print("✅ 成功點擊 icon!")
        time.sleep(0.5)

        # 再次等待元素變可點擊（確保頁面沒變動）
        icon = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//input[@type='image' and contains(@id, 'img_SignHead')]"))
        )

        icon.click()
        print("✅ 第二次點擊成功")
        time.sleep(0.5)

    except Exception as e:
        print(f"❌ 點擊 icon 失敗：{e}")

    page_number = 0
    dec = 0

    while True:
        driver.switch_to.default_content()
        driver.switch_to.frame("main")
        file_icons = driver.find_elements(By.XPATH, '//a[img[contains(@src, "../IMAGE/GDOCSIGN_1.gif")]]')
        if not file_icons:
            print("⚠️ 無電子檔")
            break
        for i, a_tag in enumerate(file_icons):
            current_odis_window = driver.current_window_handle 
            existing_windows_before_click = set(driver.window_handles)
            
            wait.until(EC.visibility_of(a_tag))
            driver.execute_script("arguments[0].click();", a_tag)

            # items_processed_count +=1 # If you want to count items like original cur_page_number
            
            doc_detail_window_handle = None
            try:
                WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(len(existing_windows_before_click) + 1))
                
                for handle_in_loop in driver.window_handles:
                    if handle_in_loop not in existing_windows_before_click:
                        doc_detail_window_handle = handle_in_loop
                        break
                
                if doc_detail_window_handle is None:
                    print(f"❌ 嚴重錯誤：雖然視窗數量增加，但無法識別文件詳細資料的新視窗句柄。跳過此文件。")
                    if driver.current_window_handle != current_odis_window:
                        driver.switch_to.window(current_odis_window)
                    driver.switch_to.default_content()
                    wait.until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "main")))
                    continue 

                driver.switch_to.window(doc_detail_window_handle)
                
                wait.until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "GDOCDetailTopMenu")))
                driver.find_element(By.XPATH, "//td[@class='MenuChild' and contains(@url, 'TableTitle=流程')]").click()

                driver.switch_to.default_content()
                wait.until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "GDOCDetailMain")))

                data_cells = driver.find_elements(By.CSS_SELECTOR, 'td.cssData')
                doc_number = data_cells[0].text.strip() if len(data_cells) > 0 else "無文號資料"
                unit = data_cells[2].text.strip() if len(data_cells) > 2 else "無單位資料"
                person = data_cells[3].text.strip() if len(data_cells) > 3 else "無人員資料"


                decision_date, archive_date, sign_date = None, None, None
                flow_rows = driver.find_elements(By.XPATH, "//tr[@class='cssGridItem' or @class='cssGridAlter']")
                for row in flow_rows:
                    cells = row.find_elements(By.TAG_NAME, "td")
                    if len(cells) >= 6:
                        date_cell_text_parts = cells[0].text.strip().split()
                        
                        if not date_cell_text_parts:
                            print(f"⚠️ 警告：文件 '{doc_number}' 的流程中日期欄位為空。跳過此流程日期處理。")
                            continue 
                        
                        process_date_str_roc_flow = date_cell_text_parts[0] # Renamed to avoid confusion
                        process_type = cells[4].text.strip()

                        try:
                            parsed_date_ad = roc_to_ad(process_date_str_roc_flow)
                            if process_type == "決行(存查)":
                                decision_date = parsed_date_ad
                            elif process_type == "存查":
                                archive_date = parsed_date_ad
                            elif process_type == "簽收":
                                sign_date = parsed_date_ad
                        except ValueError as ve:
                            print(f"⚠️ 警告：無法解析文件 '{doc_number}' 的日期 '{process_date_str_roc_flow}'。錯誤：{ve}。")
                        except Exception as e_date_conv:
                            print(f"⚠️ 警告：轉換日期時發生未知錯誤 '{process_date_str_roc_flow}' for doc '{doc_number}'. Error: {e_date_conv}")

                if decision_date and archive_date:
                    diff = working_days_diff(decision_date, archive_date)
                    if diff > 5:
                        output.append({
                            "公文文號": doc_number,
                            "主辦單位": unit,
                            "主辦人員": person,
                            "決行(存查)日期": decision_date.strftime("%Y/%m/%d"),
                            "存查日期": archive_date.strftime("%Y/%m/%d"),
                            "簽收日期": sign_date.strftime("%Y/%m/%d"),
                            "資料完整性": "✅ 完整 (符合條件)"
                        })
                        print(f"📌 {doc_number} 記錄完成 (天數差異 {diff} > 5)")

            except Exception as e_doc_processing:
                print(f"❌ 處理文件明細時發生錯誤：{e_doc_processing}")
            finally:
                current_handles_after_doc = set(driver.window_handles)
                for handle_to_close in current_handles_after_doc:
                    if handle_to_close not in existing_windows_before_click and handle_to_close != current_odis_window :
                        try:
                            driver.switch_to.window(handle_to_close)
                            driver.close()
                        except Exception as e_closing_doc_window:
                            print(f"注意：嘗試關閉文件詳細視窗時發生錯誤: {e_closing_doc_window}")
                
                driver.switch_to.window(current_odis_window)
                driver.switch_to.default_content()
                try:
                    wait.until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "main")))
                except Exception as e_main_frame_return:
                    print(f"❌ 嚴重錯誤：返回主列表 iframe 時失敗: {e_main_frame_return}。")

        # 換頁
        try:
            page_links = wait.until(EC.presence_of_all_elements_located(
                (By.XPATH, "//a[contains(@href, '__doPostBack') and contains(text(), '[')]")
            ))

            # 1st page: [ [2] , [3], ... , [ ... ] ]
            # 2nd page: [ [ ... ], [12], ... , [20], [ ... ] ]
            # ...
            # the last page: [ [ ... ], [92], ... , [100] ]

            """
            if page_links[0] == "[ ... ]" and page_links[-1] != "[ ... ]":
                # 不要去按 [ ... ], 因為已經是最後一頁
                print("⚠️ 當前頁面已經是最後一頁，無需點擊 [ ... ]")
            elif page_links[0] != "[ ... ]" and page_links[-1] == "[ ... ]":
                # 不要去按 [ ... ], 因為已經是第一頁
                print("⚠️ 當前頁面已經是第一頁，無需點擊 [ ... ]")
            else:
                # 正常情況下，第一個是 [ ... ]，最後一個是 [ ... ]
                print("✅ 頁碼按鈕正常")
            """

            # 抓出所有像 [ 1 ] 的頁碼按鈕
            page_numbers = [
                link for link in page_links
                if re.fullmatch(r"\[\s*\d+\s*\]", link.text.strip())
            ]

            # 額外抓出 [ ... ] 的按鈕 [ [ ... ] ]
            try:
                dots = driver.find_elements(By.XPATH, "//a[contains(@href, '__doPostBack') and contains(text(), '[ ... ]')]")
            except Exception as e:
                print("❌ 找不到 [ ... ] 的按鈕")
                dots = []

            print(f"目前頁碼：{page_number + 1 + dec * 10}")

            # 如果已經是第 10 頁（索引 9），點擊 [ ... ] 展開下一組頁碼
            try:
                if page_number == 9:
                    if dots:
                        dots[-1].click()
                        print("✅ 點擊 [ ... ] 展開更多頁碼")
                        page_number = 0
                        dec += 1
                    else:
                        print("⚠️ 沒有更多 [ ... ]，可能已是最後一頁")
                        break
                else:
                    # 頁碼數量不足時防呆處理
                    if page_number >= len(page_numbers):
                        print("⚠️ 頁碼索引超出範圍，停止換頁")
                        break
                    next_page = page_numbers[page_number]
                    print(f"➡️ 換頁到：{next_page.text}")
                    next_page.click()
                    page_number += 1
                time.sleep(0.1)

            except Exception as e:
                print("❌ 換頁錯誤，可能已到最後一頁")
                break

        except Exception as e:
            print(f"❌ 換頁失敗：{e}")
            break

    # 寫入 Excel
    if output:
        df = pd.DataFrame(output)
        file_path = "存查結果.xlsx"
        if not os.path.exists(file_path):
            df.to_excel(file_path, index=False, engine='openpyxl')
        else:
            try:
                # Load existing workbook to find the last row
                book = load_workbook(file_path)
                sheet_name = book.sheetnames[0] 
                start_row_index = book[sheet_name].max_row
                
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    # Make sure that the writer.book is correctly set by openpyxl when mode='a'
                    # For 'overlay', openpyxl should handle this. We need to ensure header is not written.
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row_index)
                print(f"✅ 寫入 Excel 完成：新增 {len(output)} 筆資料到 {file_path}")
            except Exception as e_excel:
                 print(f"❌ 寫入 Excel 檔案 '{file_path}' 失敗: {e_excel}")
                 timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                 fallback_path = f"存查結果_fallback_{timestamp}.xlsx"
                 df.to_excel(fallback_path, index=False, engine='openpyxl')
                 print(f"⚠️ 已將本次結果儲存到新的 fallback 檔案: {fallback_path}")
    else:
        print(f"⚠️ 月份 {start_date} ~ {end_date} 無需寫入資料")

# ========== 主流程 ==========
start_date_str, end_date_str = get_user_input()

# 將民國年轉換成西元 datetime 物件
def convert_to_date(minguo_str):
    y, m, d = map(int, minguo_str.split("/"))
    return datetime(y + 1911, m, d)

start_date = convert_to_date(start_date_str)
end_date = convert_to_date(end_date_str)

login()
original_window = driver.current_window_handle
open_odis()

current = start_date
while current <= end_date:
    start_str = f"{current.year - 1911}/{current.month:02d}/01"
    last_day = monthrange(current.year, current.month)[1]
    end_str = f"{current.year - 1911}/{current.month:02d}/{last_day}"

    print(f"\n🔍 處理 {start_str} ~ {end_str}")

    for handle in driver.window_handles:
        if handle != original_window:
            driver.switch_to.window(handle)
            break

    driver.get("https://odis.taichung.gov.tw/HOME/Home_Splitter.aspx")
    wait.until(EC.element_to_be_clickable(
        (By.XPATH, '/html/body/form/div[3]/div[1]/table/tbody/tr/td/table/tbody/tr[1]/td[3]/table'))
    ).click()

    process_month(start_str, end_str)

    # 前進一個月
    if current.month == 12:
        current = current.replace(year=current.year + 1, month=1)
    else:
        current = current.replace(month=current.month + 1)

input("📁 所有月份處理完成，按 Enter 結束")
driver.quit()